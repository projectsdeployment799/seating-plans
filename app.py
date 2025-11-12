from flask import Flask, render_template, request, send_file
import os
import sys
import tempfile
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Define base directory first
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_DIR = os.path.join(BASE_DIR, "templates")
STATIC_DIR = os.path.join(BASE_DIR, "static")

# Use temp directory for Vercel
if os.environ.get('VERCEL'):
    UPLOAD_FOLDER = tempfile.gettempdir()
    OUTPUT_FOLDER = tempfile.gettempdir()
else:
    UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")
    OUTPUT_FOLDER = os.path.join(BASE_DIR, "outputs")

app = Flask(__name__, 
            template_folder=TEMPLATE_DIR,
            static_folder=STATIC_DIR)

# Create folders if they don't exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)


# ----------- SEATING PLAN GENERATOR -----------
def create_seating_plan(input_file):
    """Generate seating plan with optimized memory usage"""
    output_path = os.path.join(OUTPUT_FOLDER, "plan_updated.xlsx")
    
    try:
        df_roll_numbers = pd.read_excel(input_file, sheet_name=0, usecols=[1])
        roll_numbers = df_roll_numbers.iloc[:, 0].dropna().tolist()
        total_roll_numbers = len(roll_numbers)

        df_rooms = pd.read_excel(input_file, sheet_name=1)
        
        # Create workbook directly without DataFrame
        wb = Workbook()
        ws = wb.active
        ws.title = 'Seating Plan'
        
        # Write headers
        headers = ['S.No', 'ROOM NO', 'ROLL NO', 'NUMBER OF STUDENTS']
        ws.append(headers)
        
        # Format headers
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="004aad", end_color="004aad", fill_type="solid")
        center_align = Alignment(horizontal='center', vertical='center')
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        ws.row_dimensions[1].height = 20
        
        # Set column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 18
        
        # Write data row by row (memory efficient)
        serial_no = 1
        current_roll_index = 0
        data_font = Font(name="Calibri", size=11)
        
        for _, room in df_rooms.iterrows():
            if current_roll_index >= total_roll_numbers:
                break
            
            room_number = room['Room Number']
            num_benches = int(room['Number of Benches'])
            start_roll = current_roll_index
            end_roll = min(current_roll_index + num_benches - 1, total_roll_numbers - 1)
            
            row_data = [
                serial_no,
                room_number,
                f"{roll_numbers[start_roll]}-{roll_numbers[end_roll]}",
                (end_roll - start_roll + 1)
            ]
            ws.append(row_data)
            
            # Format data row
            row_num = serial_no + 1
            for col_num in range(1, 5):
                cell = ws.cell(row=row_num, column=col_num)
                cell.font = data_font
                cell.alignment = center_align
                cell.border = border
            ws.row_dimensions[row_num].height = 18
            
            serial_no += 1
            current_roll_index = end_roll + 1
        
        wb.save(output_path)
        return output_path
        
    except Exception as e:
        raise Exception(f"Error creating seating plan: {str(e)}")


# ----------- DETAILED SEATING ARRANGEMENT -----------
def generate_seating_arrangement(input_file):
    """Generate detailed seating arrangement with optimized memory"""
    output_path = os.path.join(OUTPUT_FOLDER, "seating_arrangement.xlsx")
    
    try:
        df_roll_numbers = pd.read_excel(input_file, sheet_name=0, usecols=[1])
        roll_numbers = df_roll_numbers.iloc[:, 0].dropna().tolist()
        df_rooms = pd.read_excel(input_file, sheet_name=1)
        
        total_roll_numbers = len(roll_numbers)
        wb = Workbook()
        wb.remove(wb.active)
        current_roll_index = 0

        for _, room in df_rooms.iterrows():
            if current_roll_index >= total_roll_numbers:
                break

            room_number = room['Room Number']
            size = room['Size']
            rows, cols = map(int, size.split('X'))
            num_benches = int(room['Number of Benches'])

            sheet = wb.create_sheet(title=str(room_number))
            
            # Set column widths
            for i in range(1, cols * 2 + 1):
                sheet.column_dimensions[get_column_letter(i)].width = 15
            
            # Add headers
            _add_seating_headers(sheet, cols, room_number)
            
            # Add seating data
            roll_no_counter = _add_seating_data(sheet, cols, rows, current_roll_index, roll_numbers, total_roll_numbers)
            current_roll_index += (roll_no_counter - 1)
            
            # Add total row
            total_row = sheet.max_row + 2
            sheet.append(["TOTAL : " + str(roll_no_counter - 1)])
            sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=cols*2)
            total_cell = sheet.cell(row=total_row, column=1)
            total_cell.alignment = Alignment(horizontal='center', vertical='center')
            total_cell.font = Font(bold=True, size=10)
            sheet.row_dimensions[total_row].height = 18

        wb.save(output_path)
        return output_path
        
    except Exception as e:
        raise Exception(f"Error creating seating arrangement: {str(e)}")


def _add_seating_headers(sheet, cols, room_number):
    """Add headers to seating sheet"""
    sheet.append(["NARASARAOPETA ENGINEERING COLLEGE:: NARASARAOPET"])
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols*2)
    sheet.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
    sheet.cell(row=1, column=1).font = Font(bold=True, size=12)
    sheet.row_dimensions[1].height = 20
    
    sheet.append(["AUTONOMOUS"])
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols*2)
    sheet.cell(row=2, column=1).alignment = Alignment(horizontal='center', vertical='center')
    sheet.cell(row=2, column=1).font = Font(bold=True, size=11)
    sheet.row_dimensions[2].height = 18
    
    sheet.append(["MID-I EXAMINATIONS OF I B.TECH I- SEMESTER"])
    sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=cols*2)
    sheet.cell(row=3, column=1).alignment = Alignment(horizontal='center', vertical='center')
    sheet.cell(row=3, column=1).font = Font(bold=True, size=10)
    sheet.row_dimensions[3].height = 18
    
    sheet.append(["SEATING ARRANGEMENT PLAN"])
    sheet.merge_cells(start_row=4, start_column=1, end_row=4, end_column=cols*2)
    sheet.cell(row=4, column=1).alignment = Alignment(horizontal='center', vertical='center')
    sheet.cell(row=4, column=1).font = Font(bold=True, size=10)
    sheet.row_dimensions[4].height = 18
    
    sheet.append([])
    sheet.row_dimensions[5].height = 10
    
    # Room details
    total_cols = cols * 2
    section_size = total_cols // 3
    sheet.append(["", "", "", "", "", ""])
    sheet.merge_cells(start_row=6, start_column=1, end_row=6, end_column=section_size)
    sheet.cell(row=6, column=1, value=f"ROOM NO: {room_number}").font = Font(bold=True, size=10)
    sheet.cell(row=6, column=1).alignment = Alignment(horizontal='center', vertical='center')
    
    sheet.merge_cells(start_row=6, start_column=section_size+1, end_row=6, end_column=2*section_size)
    sheet.cell(row=6, column=section_size+1, value="DATE: 10/11/2025 to 14/11/2025").font = Font(bold=True, size=10)
    sheet.cell(row=6, column=section_size+1).alignment = Alignment(horizontal='center', vertical='center')
    
    sheet.merge_cells(start_row=6, start_column=(2*section_size)+1, end_row=6, end_column=total_cols)
    sheet.cell(row=6, column=(2*section_size)+1, value="MID-I SESSION: AN").font = Font(bold=True, size=10)
    sheet.cell(row=6, column=(2*section_size)+1).alignment = Alignment(horizontal='center', vertical='center')
    sheet.row_dimensions[6].height = 18
    
    sheet.append(["TIME: 02.10 PM to 04.00 PM"])
    sheet.merge_cells(start_row=7, start_column=1, end_row=7, end_column=total_cols)
    sheet.cell(row=7, column=1).font = Font(bold=True, size=10)
    sheet.cell(row=7, column=1).alignment = Alignment(horizontal='left', vertical='center')
    sheet.row_dimensions[7].height = 16
    
    sheet.append([])
    sheet.row_dimensions[8].height = 10
    
    # Row headers
    header_row = []
    for col in range(cols):
        header_row += [f"ROW-{col+1}", ""]
    sheet.append(header_row)
    for col in range(1, cols * 2, 2):
        sheet.merge_cells(start_row=9, start_column=col, end_row=9, end_column=col+1)
        sheet.cell(row=9, column=col).font = Font(bold=True, size=9)
        sheet.cell(row=9, column=col).alignment = Alignment(horizontal='center', vertical='center')
    sheet.row_dimensions[9].height = 16
    
    # Sub-headers
    sub_header = []
    for _ in range(cols):
        sub_header += ["S.NO", "I YEAR"]
    sheet.append(sub_header)
    for col in range(1, cols * 2 + 1):
        sheet.cell(row=10, column=col).font = Font(bold=True, size=9)
        sheet.cell(row=10, column=col).alignment = Alignment(horizontal='center', vertical='center')
        sheet.cell(row=10, column=col).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    sheet.row_dimensions[10].height = 16


def _add_seating_data(sheet, cols, rows, current_roll_index, roll_numbers, total_roll_numbers):
    """Add seating data to sheet (memory optimized)"""
    roll_no_counter = 1
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for r in range(rows):
        row_data = []
        for c in range(cols):
            if current_roll_index < total_roll_numbers:
                row_data += [roll_no_counter, roll_numbers[current_roll_index]]
                roll_no_counter += 1
                current_roll_index += 1
            else:
                row_data += ["", ""]
        
        sheet.append(row_data)
        row_idx = 11 + r
        for col in range(1, cols * 2 + 1):
            cell = sheet.cell(row=row_idx, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(size=9)
            cell.border = border
        sheet.row_dimensions[row_idx].height = 20
    
    return roll_no_counter


# ----------- SIGNATURE LIST -----------
def generate_signature_list(input_file):
    output_path = os.path.join(OUTPUT_FOLDER, "signature_list.xlsx")
    df_roll_numbers = pd.read_excel(input_file, sheet_name=0)
    df_rooms = pd.read_excel(input_file, sheet_name=1)
    students = df_roll_numbers.iloc[0:, [1, 2, 3, 4]].dropna()
    students.columns = ['HTNO', 'Name', 'Section', 'Branch']
    wb = Workbook()
    ws = wb.active
    ws.title = "Signature List"

    # Fonts and alignment
    regular_font = Font(name="Calibri", size=11)
    bold_font = Font(name="Calibri", bold=True, size=11)
    header_font = Font(name="Calibri", bold=True, size=12)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center')
    thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Set column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 18

    row_index = 1
    current_roll_index = 0
    total_students = len(students)

    for _, room in df_rooms.iterrows():
        if current_roll_index >= total_students:
            break

        room_number = room['Room Number']
        num_benches = int(room['Number of Benches'])
        room_students = students.iloc[current_roll_index:current_roll_index + num_benches]
        
        # College header
        ws.append(["NARASARAOPETA ENGINEERING COLLEGE : NARASARAOPET (AUTONOMOUS)"])
        ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=5)
        cell = ws.cell(row=row_index, column=1)
        cell.font = header_font
        cell.alignment = center_align
        ws.row_dimensions[row_index].height = 20
        row_index += 1
        
        # Exam details
        ws.append(["(R23) 2025 BATCH I B.TECH I SEM I MID SIGNATURE LIST NOV - 2025"])
        ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=5)
        cell = ws.cell(row=row_index, column=1)
        cell.font = header_font
        cell.alignment = center_align
        ws.row_dimensions[row_index].height = 18
        row_index += 1
        
        # Room number
        ws.append([f"Room No: {room_number}"])
        ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=5)
        cell = ws.cell(row=row_index, column=1)
        cell.font = bold_font
        cell.alignment = left_align
        ws.row_dimensions[row_index].height = 16
        row_index += 1
        
        # Empty row
        ws.append([])
        ws.row_dimensions[row_index].height = 8
        row_index += 1
        
        # Column headers
        ws.append(["S.NO", "HTNO", "Name", "Answer Booklet No", "Signature"])
        for col in range(1, 6):
            cell = ws.cell(row=row_index, column=col)
            cell.font = bold_font
            cell.alignment = center_align
            cell.border = thin
        ws.row_dimensions[row_index].height = 18
        row_index += 1

        # Student entries
        for i, (htno, name) in enumerate(room_students[['HTNO', 'Name']].values, start=1):
            ws.append([i, htno, name, "", ""])
            for col in range(1, 6):
                cell = ws.cell(row=row_index, column=col)
                cell.font = regular_font
                cell.border = thin
                if col == 1:
                    cell.alignment = center_align
                elif col in [2, 4, 5]:
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align
            ws.row_dimensions[row_index].height = 20
            row_index += 1

        # Empty rows between rooms
        ws.append([])
        ws.row_dimensions[row_index].height = 12
        row_index += 1
        ws.append([])
        ws.row_dimensions[row_index].height = 12
        row_index += 1
        
        current_roll_index += num_benches

    wb.save(output_path)
    return output_path


def cleanup_files():
    """Remove all files from upload and output directories"""
    try:
        for filename in os.listdir(UPLOAD_FOLDER):
            file_path = os.path.join(UPLOAD_FOLDER, filename)
            if os.path.isfile(file_path):
                os.remove(file_path)
        
        for filename in os.listdir(OUTPUT_FOLDER):
            file_path = os.path.join(OUTPUT_FOLDER, filename)
            if os.path.isfile(file_path):
                os.remove(file_path)
    except Exception as e:
        print(f"Cleanup error: {str(e)}")


@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        if 'file' not in request.files:
            return render_template("index.html", error="No file provided")
        
        file = request.files['file']
        if file.filename == '':
            return render_template("index.html", error="No file selected")
        
        if not file.filename.endswith('.xlsx'):
            return render_template("index.html", error="Please upload an Excel file")
        
        try:
            filepath = os.path.join(UPLOAD_FOLDER, file.filename)
            file.save(filepath)
            
            # Generate files
            plan_path = create_seating_plan(filepath)
            arrangement_path = generate_seating_arrangement(filepath)
            signature_path = generate_signature_list(filepath)
            
            # Return with correct variable names
            return render_template("index.html",
                                 plan_file='plan_updated.xlsx',
                                 arrangement_file='seating_arrangement.xlsx',
                                 signature_file='signature_list.xlsx')
        except Exception as e:
            return render_template("index.html", error=f"Error: {str(e)}")
    
    return render_template("index.html")


@app.route("/download/<filename>")
def download(filename):
    try:
        file_path = os.path.join(OUTPUT_FOLDER, filename)
        if not os.path.exists(file_path):
            return "File not found", 404
        
        return send_file(file_path, as_attachment=True)
    except Exception as e:
        return f"Error: {str(e)}", 500


@app.route("/reset")
def reset():
    try:
        cleanup_files()
        return {"status": "success"}
    except Exception as e:
        return {"status": "error", "message": str(e)}, 500


if __name__ == "__main__":
    # Only enable debug locally
    is_local = not os.environ.get('VERCEL')
    app.run(debug=is_local)
